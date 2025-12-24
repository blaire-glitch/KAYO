from functools import wraps
from io import BytesIO
from datetime import datetime
from flask import Blueprint, render_template, redirect, url_for, flash, request, Response, current_app, jsonify
from flask_login import login_required, current_user
from app import db
from app.models.user import User
from app.models.delegate import Delegate
from app.models.payment import Payment
from app.models.operations import CheckInRecord
from app.models.fund_management import Pledge, ScheduledPayment, FundTransfer, FundTransferApproval, PaymentSummary
from app.models.permission_request import PermissionRequest
from app.forms import AdminUserForm, SearchForm, CheckInForm
from sqlalchemy import text

admin_bp = Blueprint('admin', __name__, url_prefix='/admin')


def admin_required(f):
    """Decorator to require admin role"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or not current_user.is_admin():
            flash('You do not have permission to access this page.', 'danger')
            return redirect(url_for('main.dashboard'))
        return f(*args, **kwargs)
    return decorated_function


@admin_bp.route('/')
@login_required
@admin_required
def dashboard():
    """Admin dashboard with overview"""
    # Get overall stats
    total_delegates = Delegate.query.count()
    paid_delegates = Delegate.query.filter_by(is_paid=True).count()
    unpaid_delegates = total_delegates - paid_delegates
    checked_in = Delegate.query.filter_by(checked_in=True).count()
    
    total_users = User.query.filter(User.role != 'admin').count()
    total_collected = Payment.get_total_collected()
    
    # Get stats by archdeaconry
    archdeaconry_stats = Delegate.get_stats_by_archdeaconry()
    
    # Get stats by parish
    parish_stats = Delegate.get_stats_by_parish()
    
    # Get gender stats
    gender_stats = Delegate.get_gender_stats()
    
    # Get category stats - convert to JSON-serializable format
    category_stats_raw = Delegate.get_category_stats()
    category_stats = [{'category': row.category or 'Unknown', 'count': row.count} for row in category_stats_raw]
    
    # Get daily registration stats (last 30 days) - convert to JSON-serializable format
    daily_stats_raw = Delegate.get_daily_registration_stats(30)
    daily_stats = [{'date': str(row.date), 'count': row.count} for row in daily_stats_raw]
    
    # Recent payments
    recent_payments = Payment.query.filter_by(
        status='completed'
    ).order_by(Payment.completed_at.desc()).limit(10).all()
    
    # Recent registrations
    recent_delegates = Delegate.query.order_by(
        Delegate.registered_at.desc()
    ).limit(10).all()
    
    return render_template('admin/dashboard.html',
        total_delegates=total_delegates,
        paid_delegates=paid_delegates,
        unpaid_delegates=unpaid_delegates,
        checked_in=checked_in,
        total_users=total_users,
        total_collected=total_collected,
        archdeaconry_stats=archdeaconry_stats,
        parish_stats=parish_stats,
        gender_stats=gender_stats,
        category_stats=category_stats,
        daily_stats=daily_stats,
        recent_payments=recent_payments,
        recent_delegates=recent_delegates
    )


@admin_bp.route('/delegates')
@login_required
@admin_required
def all_delegates():
    """View all delegates with filters"""
    page = request.args.get('page', 1, type=int)
    per_page = 50
    
    # Filter parameters
    archdeaconry = request.args.get('archdeaconry', '')
    parish = request.args.get('parish', '')
    payment_status = request.args.get('payment_status', '')
    gender = request.args.get('gender', '')
    search = request.args.get('search', '')
    
    # Build query
    query = Delegate.query
    
    if archdeaconry:
        query = query.filter(Delegate.archdeaconry == archdeaconry)
    if parish:
        query = query.filter(Delegate.parish == parish)
    if payment_status == 'paid':
        query = query.filter(Delegate.is_paid == True)
    elif payment_status == 'unpaid':
        query = query.filter(Delegate.is_paid == False)
    if gender:
        query = query.filter(Delegate.gender == gender)
    if search:
        query = query.filter(Delegate.name.ilike(f'%{search}%'))
    
    delegates = query.order_by(Delegate.registered_at.desc()).paginate(
        page=page, per_page=per_page, error_out=False
    )
    
    # Get unique values for filters
    archdeaconries = db.session.query(Delegate.archdeaconry).distinct().all()
    parishes = db.session.query(Delegate.parish).distinct().all()
    
    return render_template('admin/delegates.html',
        delegates=delegates,
        archdeaconries=[a[0] for a in archdeaconries],
        parishes=[p[0] for p in parishes],
        filters={
            'archdeaconry': archdeaconry,
            'parish': parish,
            'payment_status': payment_status,
            'gender': gender,
            'search': search
        }
    )


@admin_bp.route('/users')
@login_required
@admin_required
def all_users():
    """View all users"""
    page = request.args.get('page', 1, type=int)
    per_page = 50
    
    users = User.query.order_by(User.created_at.desc()).paginate(
        page=page, per_page=per_page, error_out=False
    )
    
    return render_template('admin/users.html', users=users)


@admin_bp.route('/delegates/<int:id>/delete', methods=['POST'])
@login_required
@admin_required
def admin_delete_delegate(id):
    """Delete a delegate from admin panel"""
    delegate = Delegate.query.get_or_404(id)
    
    try:
        delegate_name = delegate.name
        ticket_number = delegate.ticket_number
        
        # Unlink from payment (don't delete payment as it may cover other delegates)
        delegate.payment_id = None
        
        # Delete check-in records
        CheckInRecord.query.filter_by(delegate_id=id).delete()
        
        # Delete the delegate
        db.session.delete(delegate)
        db.session.commit()
        
        flash(f'Delegate "{delegate_name}" ({ticket_number}) deleted successfully.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting delegate: {str(e)}', 'danger')
    
    return redirect(url_for('admin.all_delegates'))


@admin_bp.route('/delegates/bulk-delete', methods=['POST'])
@login_required
@admin_required
def bulk_delete_delegates():
    """Bulk delete delegates from admin panel"""
    delegate_ids = request.form.getlist('delegate_ids', type=int)
    
    if not delegate_ids:
        flash('No delegates selected for deletion.', 'warning')
        return redirect(url_for('admin.all_delegates'))
    
    try:
        deleted_count = 0
        skipped_count = 0
        
        for delegate_id in delegate_ids:
            delegate = Delegate.query.get(delegate_id)
            if delegate:
                # Unlink from payment (don't delete payment as it may cover other delegates)
                delegate.payment_id = None
                
                # Delete check-in records
                CheckInRecord.query.filter_by(delegate_id=delegate_id).delete()
                
                # Delete the delegate
                db.session.delete(delegate)
                deleted_count += 1
            else:
                skipped_count += 1
        
        db.session.commit()
        
        if deleted_count > 0:
            flash(f'Successfully deleted {deleted_count} delegate(s).', 'success')
        if skipped_count > 0:
            flash(f'{skipped_count} delegate(s) were not found and skipped.', 'info')
            
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting delegates: {str(e)}', 'danger')
    
    return redirect(url_for('admin.all_delegates'))


@admin_bp.route('/users/create', methods=['GET', 'POST'])
@login_required
@admin_required
def create_user():
    """Create a new user"""
    form = AdminUserForm()
    
    if form.validate_on_submit():
        # Check if email already exists
        if User.query.filter_by(email=form.email.data).first():
            flash('Email already registered.', 'danger')
            return render_template('admin/user_form.html', form=form, action='Create')
        
        user = User(
            name=form.name.data,
            email=form.email.data,
            phone=form.phone.data or None,
            role=form.role.data,
            local_church=form.local_church.data or None,
            parish=form.parish.data or None,
            archdeaconry=form.archdeaconry.data or None,
            is_approved=True,
            approval_status='approved',
            approved_by=current_user.id,
            approved_at=datetime.utcnow()
        )
        user.set_password(form.password.data)
        db.session.add(user)
        db.session.commit()
        flash(f'User "{user.name}" created successfully!', 'success')
        return redirect(url_for('admin.all_users'))
    
    return render_template('admin/user_form.html', form=form, action='Create')


@admin_bp.route('/users/<int:id>/edit', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_user(id):
    """Edit a user"""
    user = User.query.get_or_404(id)
    form = AdminUserForm(obj=user)
    
    if form.validate_on_submit():
        try:
            # Check if email changed and already exists
            if form.email.data != user.email:
                existing_user = User.query.filter_by(email=form.email.data).first()
                if existing_user:
                    flash('Email already registered.', 'danger')
                    return render_template('admin/user_form.html', form=form, action='Edit', user=user)
            
            # Check if phone changed and already exists
            if form.phone.data and form.phone.data != user.phone:
                existing_phone = User.query.filter_by(phone=form.phone.data).first()
                if existing_phone and existing_phone.id != user.id:
                    flash('Phone number already registered.', 'danger')
                    return render_template('admin/user_form.html', form=form, action='Edit', user=user)
            
            user.name = form.name.data
            user.email = form.email.data
            user.phone = form.phone.data or None
            user.role = form.role.data
            user.local_church = form.local_church.data or None
            user.parish = form.parish.data or None
            user.archdeaconry = form.archdeaconry.data or None
            
            if form.password.data:
                user.set_password(form.password.data)
            
            db.session.commit()
            flash(f'User "{user.name}" updated successfully!', 'success')
            return redirect(url_for('admin.all_users'))
        except Exception as e:
            db.session.rollback()
            current_app.logger.error(f'Error editing user {id}: {str(e)}')
            error_message = str(e).lower()
            if 'unique constraint' in error_message and 'phone' in error_message:
                flash('Phone number already registered by another user.', 'danger')
            elif 'unique constraint' in error_message and 'email' in error_message:
                flash('Email already registered by another user.', 'danger')
            else:
                flash(f'An error occurred while updating the user: {str(e)}', 'danger')
            return render_template('admin/user_form.html', form=form, action='Edit', user=user)
    
    return render_template('admin/user_form.html', form=form, action='Edit', user=user)


@admin_bp.route('/users/<int:id>/toggle-active', methods=['POST'])
@login_required
@admin_required
def toggle_user_active(id):
    """Toggle user active status"""
    user = User.query.get_or_404(id)
    
    # Don't allow deactivating yourself
    if user.id == current_user.id:
        flash('You cannot deactivate your own account.', 'danger')
        return redirect(url_for('admin.all_users'))
    
    user.is_active = not user.is_active
    db.session.commit()
    
    status = 'activated' if user.is_active else 'deactivated'
    flash(f'User "{user.name}" has been {status}.', 'success')
    return redirect(url_for('admin.all_users'))


@admin_bp.route('/users/<int:id>/delete', methods=['POST'])
@login_required
@admin_required
def delete_user(id):
    """Permanently delete a user and optionally their delegates"""
    user = User.query.get_or_404(id)
    
    # Don't allow deleting yourself
    if user.id == current_user.id:
        flash('You cannot delete your own account.', 'danger')
        return redirect(url_for('admin.all_users'))
    
    # Don't allow deleting super_admin
    if user.role == 'super_admin':
        flash('Super admin accounts cannot be deleted.', 'danger')
        return redirect(url_for('admin.all_users'))
    
    try:
        user_name = user.name
        user_email = user.email
        delegate_count = user.delegates.count()
        
        # Check if we should delete delegates too
        delete_delegates = request.form.get('delete_delegates') == 'yes'
        
        # Handle fund transfers - reassign to admin or delete
        FundTransfer.query.filter_by(from_user_id=user.id).update({'from_user_id': current_user.id})
        FundTransfer.query.filter_by(to_user_id=user.id).update({'to_user_id': current_user.id})
        
        # Handle fund transfer approvals
        FundTransferApproval.query.filter_by(approved_by=user.id).update({'approved_by': current_user.id})
        
        # Handle pledges - reassign recorded_by
        Pledge.query.filter_by(recorded_by=user.id).update({'recorded_by': current_user.id})
        
        # Handle scheduled payments - reassign recorded_by
        ScheduledPayment.query.filter_by(recorded_by=user.id).update({'recorded_by': current_user.id})
        
        # Handle payment summaries
        PaymentSummary.query.filter_by(user_id=user.id).update({'user_id': current_user.id})
        
        # Handle permission requests - delete them
        PermissionRequest.query.filter_by(user_id=user.id).delete()
        
        # Handle payments - reassign to admin
        Payment.query.filter_by(user_id=user.id).update({'user_id': current_user.id})
        
        if delete_delegates and delegate_count > 0:
            # Delete all delegates registered by this user
            for delegate in user.delegates.all():
                # Delete check-in records first
                CheckInRecord.query.filter_by(delegate_id=delegate.id).delete()
                db.session.delete(delegate)
        elif delegate_count > 0:
            # Reassign delegates to admin
            for delegate in user.delegates.all():
                delegate.registered_by = current_user.id
        
        # Delete the user
        db.session.delete(user)
        db.session.commit()
        
        if delete_delegates:
            flash(f'User "{user_name}" ({user_email}) and their {delegate_count} delegates deleted successfully.', 'success')
        else:
            flash(f'User "{user_name}" ({user_email}) deleted successfully. {delegate_count} delegates were reassigned to you.', 'success')
            
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f'Error deleting user {id}: {str(e)}')
        flash(f'Error deleting user: {str(e)}', 'danger')
    
    return redirect(url_for('admin.all_users'))


@admin_bp.route('/users/delete-inactive', methods=['POST'])
@login_required
@admin_required
def delete_inactive_users():
    """Delete all inactive users (bulk action)"""
    try:
        # Get all inactive users except admins and super_admins
        inactive_users = User.query.filter(
            User.is_active == False,
            User.role.notin_(['admin', 'super_admin']),
            User.id != current_user.id
        ).all()
        
        if not inactive_users:
            flash('No inactive users to delete.', 'info')
            return redirect(url_for('admin.all_users'))
        
        deleted_count = 0
        delegate_count = 0
        
        for user in inactive_users:
            user_delegates = user.delegates.count()
            
            # Handle fund transfers - reassign to admin
            FundTransfer.query.filter_by(from_user_id=user.id).update({'from_user_id': current_user.id})
            FundTransfer.query.filter_by(to_user_id=user.id).update({'to_user_id': current_user.id})
            
            # Handle fund transfer approvals
            FundTransferApproval.query.filter_by(approved_by=user.id).update({'approved_by': current_user.id})
            
            # Handle pledges
            Pledge.query.filter_by(recorded_by=user.id).update({'recorded_by': current_user.id})
            
            # Handle scheduled payments
            ScheduledPayment.query.filter_by(recorded_by=user.id).update({'recorded_by': current_user.id})
            
            # Handle payment summaries
            PaymentSummary.query.filter_by(user_id=user.id).update({'user_id': current_user.id})
            
            # Handle permission requests
            PermissionRequest.query.filter_by(user_id=user.id).delete()
            
            # Handle payments
            Payment.query.filter_by(user_id=user.id).update({'user_id': current_user.id})
            
            # Delete user's delegates
            for delegate in user.delegates.all():
                CheckInRecord.query.filter_by(delegate_id=delegate.id).delete()
                db.session.delete(delegate)
                delegate_count += 1
            
            db.session.delete(user)
            deleted_count += 1
        
        db.session.commit()
        flash(f'Deleted {deleted_count} inactive users and {delegate_count} associated delegates.', 'success')
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f'Error deleting inactive users: {str(e)}')
        flash(f'Error deleting inactive users: {str(e)}', 'danger')
    
    return redirect(url_for('admin.all_users'))


@admin_bp.route('/payments')
@login_required
@admin_required
def all_payments():
    """View all payments"""
    page = request.args.get('page', 1, type=int)
    per_page = 50
    
    status = request.args.get('status', '')
    
    query = Payment.query
    if status:
        query = query.filter(Payment.status == status)
    
    payments = query.order_by(Payment.created_at.desc()).paginate(
        page=page, per_page=per_page, error_out=False
    )
    
    # Payment stats
    payment_stats = Payment.get_payment_stats()
    
    return render_template('admin/payments.html',
        payments=payments,
        payment_stats=payment_stats,
        status_filter=status
    )


@admin_bp.route('/export/delegates')
@login_required
@admin_required
def export_delegates_excel():
    """Export delegates to Excel"""
    try:
        from openpyxl import Workbook
        
        # Get filter parameters
        archdeaconry = request.args.get('archdeaconry', '')
        parish = request.args.get('parish', '')
        payment_status = request.args.get('payment_status', '')
        
        # Build query
        query = Delegate.query
        
        if archdeaconry:
            query = query.filter(Delegate.archdeaconry == archdeaconry)
        if parish:
            query = query.filter(Delegate.parish == parish)
        if payment_status == 'paid':
            query = query.filter(Delegate.is_paid == True)
        elif payment_status == 'unpaid':
            query = query.filter(Delegate.is_paid == False)
        
        delegates = query.order_by(Delegate.archdeaconry, Delegate.parish, Delegate.name).all()
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Delegates"
        
        # Headers
        headers = ['No.', 'Name', 'Gender', 'Local Church', 'Parish', 'Archdeaconry', 'Phone', 'Payment Status', 'Registered By', 'Date Registered']
        ws.append(headers)
        
        # Style headers
        for col in range(1, len(headers) + 1):
            ws.cell(row=1, column=col).font = ws.cell(row=1, column=col).font.copy(bold=True)
        
        # Data
        for idx, delegate in enumerate(delegates, 1):
            ws.append([
                idx,
                delegate.name,
                delegate.gender.capitalize(),
                delegate.local_church,
                delegate.parish,
                delegate.archdeaconry,
                delegate.phone_number or 'N/A',
                'Paid' if delegate.is_paid else 'Unpaid',
                delegate.registered_by_user.name if delegate.registered_by_user else 'N/A',
                delegate.registered_at.strftime('%Y-%m-%d %H:%M')
            ])
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return Response(
            output.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': 'attachment;filename=delegates_export.xlsx'}
        )
    except ImportError:
        flash('Excel export requires openpyxl library.', 'danger')
        return redirect(url_for('admin.all_delegates'))


@admin_bp.route('/export/delegates/pdf')
@login_required
@admin_required
def export_delegates_pdf():
    """Export delegates to PDF"""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
        from reportlab.lib.styles import getSampleStyleSheet
        
        # Get filter parameters
        archdeaconry = request.args.get('archdeaconry', '')
        parish = request.args.get('parish', '')
        payment_status = request.args.get('payment_status', '')
        
        # Build query
        query = Delegate.query
        
        if archdeaconry:
            query = query.filter(Delegate.archdeaconry == archdeaconry)
        if parish:
            query = query.filter(Delegate.parish == parish)
        if payment_status == 'paid':
            query = query.filter(Delegate.is_paid == True)
        elif payment_status == 'unpaid':
            query = query.filter(Delegate.is_paid == False)
        
        delegates = query.order_by(Delegate.archdeaconry, Delegate.parish, Delegate.name).all()
        
        # Create PDF
        output = BytesIO()
        doc = SimpleDocTemplate(output, pagesize=landscape(A4))
        
        styles = getSampleStyleSheet()
        elements = []
        
        # Title
        elements.append(Paragraph("KAYO Delegates Report", styles['Heading1']))
        
        # Table data
        data = [['No.', 'Name', 'Gender', 'Parish', 'Archdeaconry', 'Status']]
        
        for idx, delegate in enumerate(delegates, 1):
            data.append([
                str(idx),
                delegate.name,
                delegate.gender.capitalize(),
                delegate.parish,
                delegate.archdeaconry,
                'Paid' if delegate.is_paid else 'Unpaid'
            ])
        
        # Create table
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(table)
        doc.build(elements)
        
        output.seek(0)
        
        return Response(
            output.getvalue(),
            mimetype='application/pdf',
            headers={'Content-Disposition': 'attachment;filename=delegates_report.pdf'}
        )
    except ImportError:
        flash('PDF export requires reportlab library.', 'danger')
        return redirect(url_for('admin.all_delegates'))


@admin_bp.route('/check-in', methods=['GET', 'POST'])
@login_required
@admin_required
def check_in():
    """Check-in delegates using ticket number or QR scan"""
    form = CheckInForm()
    delegate = None
    
    if form.validate_on_submit():
        ticket_number = form.ticket_number.data.strip().upper()
        delegate = Delegate.query.filter_by(ticket_number=ticket_number).first()
        
        if delegate:
            if delegate.checked_in:
                flash(f'{delegate.name} is already checked in!', 'warning')
            else:
                delegate.checked_in = True
                delegate.checked_in_at = datetime.utcnow()
                db.session.commit()
                flash(f'{delegate.name} successfully checked in!', 'success')
        else:
            flash(f'No delegate found with ticket: {ticket_number}', 'danger')
    
    # Recent check-ins
    recent_checkins = Delegate.query.filter_by(
        checked_in=True
    ).order_by(Delegate.checked_in_at.desc()).limit(20).all()
    
    # Stats
    total_checked_in = Delegate.query.filter_by(checked_in=True).count()
    total_delegates = Delegate.query.count()
    
    return render_template('admin/check_in.html', 
        form=form, 
        delegate=delegate,
        recent_checkins=recent_checkins,
        total_checked_in=total_checked_in,
        total_delegates=total_delegates
    )


@admin_bp.route('/search')
@login_required
@admin_required
def search():
    """Smart search for delegates"""
    form = SearchForm(request.args)
    delegates = []
    
    query = request.args.get('query', '').strip()
    archdeaconry = request.args.get('archdeaconry', '')
    parish = request.args.get('parish', '')
    gender = request.args.get('gender', '')
    payment_status = request.args.get('payment_status', '')
    category = request.args.get('category', '')
    
    if query or archdeaconry or parish or gender or payment_status or category:
        q = Delegate.query
        
        if query:
            search_term = f"%{query}%"
            q = q.filter(
                db.or_(
                    Delegate.name.ilike(search_term),
                    Delegate.phone_number.ilike(search_term),
                    Delegate.id_number.ilike(search_term),
                    Delegate.ticket_number.ilike(search_term),
                    Delegate.local_church.ilike(search_term)
                )
            )
        
        if archdeaconry:
            q = q.filter(Delegate.archdeaconry == archdeaconry)
        if parish:
            q = q.filter(Delegate.parish == parish)
        if gender:
            q = q.filter(Delegate.gender == gender)
        if category:
            q = q.filter(Delegate.category == category)
        if payment_status == 'paid':
            q = q.filter(Delegate.is_paid == True)
        elif payment_status == 'unpaid':
            q = q.filter(Delegate.is_paid == False)
        
        delegates = q.order_by(Delegate.name).limit(100).all()
    
    return render_template('admin/search.html', form=form, delegates=delegates)


@admin_bp.route('/api/stats')
@login_required
@admin_required
def api_stats():
    """API endpoint for dashboard statistics"""
    daily_stats = Delegate.get_daily_registration_stats(30)
    
    return jsonify({
        'daily_registrations': [
            {'date': str(stat.date), 'count': stat.count} 
            for stat in daily_stats
        ],
        'total_delegates': Delegate.query.count(),
        'paid_delegates': Delegate.query.filter_by(is_paid=True).count(),
        'checked_in': Delegate.query.filter_by(checked_in=True).count()
    })


@admin_bp.route('/pending-approvals')
@login_required
@admin_required
def pending_approvals():
    """View pending user registration approvals"""
    pending_users = User.get_pending_registrations()
    return render_template('admin/pending_approvals.html', pending_users=pending_users)


@admin_bp.route('/approve-user/<int:user_id>', methods=['POST'])
@login_required
@admin_required
def approve_user(user_id):
    """Approve a pending user registration"""
    user = User.query.get_or_404(user_id)
    
    if user.approval_status != 'pending':
        flash('This registration has already been processed.', 'warning')
        return redirect(url_for('admin.pending_approvals'))
    
    # Check again if parish already has a chair
    if user.role == 'chair' and User.parish_has_chair(user.parish):
        existing_chair = User.get_parish_chair(user.parish)
        flash(f'Cannot approve. Parish "{user.parish}" already has an approved chair ({existing_chair.name}).', 'danger')
        return redirect(url_for('admin.pending_approvals'))
    
    user.is_approved = True
    user.approval_status = 'approved'
    user.approved_by = current_user.id
    user.approved_at = datetime.utcnow()
    db.session.commit()
    
    flash(f'User {user.name} has been approved as chair for {user.parish}.', 'success')
    return redirect(url_for('admin.pending_approvals'))


@admin_bp.route('/reject-user/<int:user_id>', methods=['POST'])
@login_required
@admin_required
def reject_user(user_id):
    """Reject a pending user registration"""
    user = User.query.get_or_404(user_id)
    
    if user.approval_status != 'pending':
        flash('This registration has already been processed.', 'warning')
        return redirect(url_for('admin.pending_approvals'))
    
    reason = request.form.get('reason', 'Not specified')
    
    user.is_approved = False
    user.approval_status = 'rejected'
    user.rejection_reason = reason
    user.approved_by = current_user.id
    user.approved_at = datetime.utcnow()
    db.session.commit()
    
    flash(f'User {user.name} registration has been rejected.', 'info')
    return redirect(url_for('admin.pending_approvals'))
